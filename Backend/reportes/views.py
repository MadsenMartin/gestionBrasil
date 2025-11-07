from tesoreria.models import Registro, DolarMEP, Presupuesto
from tesoreria.serializers import RegistroSerializer, RegistroCC, PresupuestoViewSerializer, RegistroListSerializer
from iva.models import Persona, ClienteProyecto
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.http import HttpResponse
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from django.db.models import OuterRef, Subquery, Value, DecimalField, Sum, F, Window
from django.db.models.functions import Coalesce
from decimal import Decimal

EXCLUDED_TIPO_REGS = ["FCV", "REC", "ISF", "RETH", "OP"]

class MDOVSPresupuesto(APIView):
    def get(self, request):
        presupuesto_id = request.query_params.get('presupuesto_id')
        if not presupuesto_id:
            return Response({'detail': 'Debe especificar un presupuesto'}, status=status.HTTP_400_BAD_REQUEST)
        presupuesto = Presupuesto.objects.get(id=presupuesto_id)
        registros = Registro.objects.filter(presupuesto=presupuesto)
        return mdo_vs_ppto_data([presupuesto], registros)

class GastosPorCasa(APIView):
    '''
    Vista que devuelve un listado de los gastos por casa, destinado a la carga de "cuentas corrientes" de los inversores.
    '''
    def get(self, request):
        cliente_proyecto_id = request.query_params.get('cliente_proyecto_id')
        if not cliente_proyecto_id:
            return Response({'detail': 'Debe especificar un cliente/proyecto'}, status=status.HTTP_400_BAD_REQUEST)
        registros = Registro.objects.filter(cliente_proyecto__id=cliente_proyecto_id)

        fechas = [x['fecha_reg'] for x in registros.values('fecha_reg')]
        # Obtener todos los DolarMEP en un solo query
        dolar_mep_values = DolarMEP.objects.filter(fecha__in=fechas).values('fecha', 'compra')

        # Crear un diccionario: {fecha: compra}
        dolar_map = {str(item['fecha']): float(item['compra']) for item in dolar_mep_values}
        table = []

        for registro in registros:
            total_gasto_ingreso = (registro.monto_gasto_ingreso_neto if registro.monto_gasto_ingreso_neto else Decimal(0)) + (registro.iva_gasto_ingreso if registro.iva_gasto_ingreso else Decimal(0))
            fecha_str = str(registro.fecha_reg)
            row = {
                "Fecha Reg": registro.fecha_reg,
                "Tipo Doc": registro.tipo_reg,
                "Proveedor": registro.proveedor.nombre_fantasia if registro.proveedor else "",
                "Imputación": registro.imputacion.imputacion if registro.imputacion else "",
                "Observaciones": registro.observacion,
                "ID_Presupuesto": registro.presupuesto.__str__() if registro.presupuesto else "",
                "Total Gasto/Ingreso": total_gasto_ingreso,
                "Total Gasto/Ingreso USD": total_gasto_ingreso / Decimal(dolar_map[fecha_str]) if fecha_str in dolar_map else 0,
            }
            table.append(row)

        if request.query_params.get('export', None) == 'true':
            return self.export_to_excel(table, "casa")

        return Response(table)
    
    def export_to_excel(self, data, entidad=None):
        try:
            if not data:
                return Response("No data available to export", status=status.HTTP_400_BAD_REQUEST)
            if not entidad:
                entidad = data[0][self.entidad].replace(" ","_")
            df = pd.DataFrame(data)

            # Create the HttpResponse object with appropriate headers
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Gastos_{entidad}.xlsx'

            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Write DataFrame to the worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create a table in the worksheet
            tab = Table(displayName=f"Gastos_{entidad}", ref=ws.dimensions)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style

            # Add the table to the worksheet
            ws.add_table(tab)

            # Save the workbook to the response
            wb.save(response)
            return response

        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

class MDOVSPresupuestoPorProveedor(APIView):
    def get(self, request):
        proveedor_id = request.query_params.get('proveedor_id')
        cliente_proyecto_id = request.query_params.get('cliente_proyecto_id')
        if not proveedor_id and not cliente_proyecto_id:
            return Response({'detail': 'Debe especificar un proveedor o cliente/proyecto'}, status=status.HTTP_400_BAD_REQUEST)
        if cliente_proyecto_id:
            presupuestos = Presupuesto.objects.filter(cliente_proyecto__id=cliente_proyecto_id)
            registros = Registro.objects.filter(cliente_proyecto__id=cliente_proyecto_id, presupuesto__isnull=False)
        else:
            presupuestos = Presupuesto.objects.filter(proveedor__id=proveedor_id)
            registros = Registro.objects.filter(proveedor__id=proveedor_id, presupuesto__isnull=False)
        export = request.query_params.get('export', None)
        if export == 'true':
            return self.export_to_excel(mdo_vs_ppto_data(presupuestos, registros))
        return Response(mdo_vs_ppto_data(presupuestos, registros))
    
    def export_to_excel(self, data):
        try:
            if not data:
                return Response("No data available to export", status=status.HTTP_400_BAD_REQUEST)
            df = pd.DataFrame(data)

            # Create the HttpResponse object with appropriate headers
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            proveedor = data[0]["Proveedor"].replace(" ","_")
            response['Content-Disposition'] = f'attachment; filename=MDO_vs_PPTO_{proveedor}.xlsx'

            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Write DataFrame to the worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create a table in the worksheet
            tab = Table(displayName=f"MDO_vs_PPTO_{proveedor}", ref=ws.dimensions)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style

            # Add the table to the worksheet
            ws.add_table(tab)

            # Save the workbook to the response
            wb.save(response)
            return response
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

def mdo_vs_ppto_data(presupuestos: list[Presupuesto], registros: list[Registro]) -> Response:
    rows = []
    for presupuesto in presupuestos:
        presupuesto_row= {
            "ID": presupuesto.id,
            "Fecha": presupuesto.fecha,
            "Tipo reg": "PPTO",
            "Cliente/Proyecto": presupuesto.cliente_proyecto.cliente_proyecto,
            "Proveedor": presupuesto.proveedor.nombre_fantasia,
            "Observaciones": presupuesto.observacion,
            "Presupuesto": presupuesto.__str__(),
            "Saldo (Monto OP/REC)": presupuesto.monto,
            "Monto": presupuesto.monto,
            "Caja": "",
        }
        rows.append(presupuesto_row)
    for registro in registros:
        row = {
            "ID": registro.id,
            "Fecha": registro.fecha_reg,
            "Tipo reg": registro.tipo_reg,
            "Cliente/Proyecto": registro.cliente_proyecto.cliente_proyecto,
            "Proveedor": registro.proveedor.nombre_fantasia,
            "Observaciones": registro.observacion,
            "Presupuesto": registro.presupuesto.__str__(),
            "Saldo (Monto OP/REC)": (registro.monto_gasto_ingreso_neto if registro.monto_gasto_ingreso_neto else Decimal(0)) + (registro.iva_gasto_ingreso if registro.iva_gasto_ingreso else Decimal(0)),
            "Caja": registro.caja.caja,
        }
        rows.append(row)
    return(rows)

class PresupuestosPorEntidad(APIView):
    serializer_class = PresupuestoViewSerializer
    queryset = Presupuesto.objects.all()
    entidad = None

    def get(self, request):
        presupuestos = self.get_filtered_presupuestos(request)
        if isinstance(presupuestos, Response):
            return presupuestos
        
        presupuestos = PresupuestoViewSerializer(presupuestos, many=True).data

        presupuestos = [{**x, 'monto': float(x.get('monto') or 0)} for x in presupuestos]
        presupuestos = [{**x, 'saldo': float(x.get('saldo') or 0)} for x in presupuestos]

        if request.query_params.get('export', None) == 'true':
            return self.export_to_excel(presupuestos)

        return Response(presupuestos)
    
    def get_filtered_presupuestos(self, request):
        presupuestos = Presupuesto.objects.all()

        # Filtro por año y mes
        añomes_max = request.query_params.get('anomes_max', None)
        añomes_min = request.query_params.get('anomes_min', None)

        if añomes_max not in [None, ""]:
            presupuestos = presupuestos.filter(fecha__lte=añomes_max)
        if añomes_min not in [None, ""]:
            presupuestos = presupuestos.filter(fecha__gte=añomes_min)

        # Filtro por entidad
        if self.entidad:
            entidad_id = request.query_params.get(self.entidad, None)
            if entidad_id is not None:
                return presupuestos.filter(**{self.entidad: entidad_id})
            else:
                return Response(f"Debe especificar un {self.entidad} válido", status=status.HTTP_400_BAD_REQUEST)
    
    def export_to_excel(self, data):
        try:
            if not data:
                return Response("No data available to export", status=status.HTTP_400_BAD_REQUEST)
            df = pd.DataFrame(data)

            # Create the HttpResponse object with appropriate headers
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Presupuestos_{self.entidad}.xlsx'

            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Write DataFrame to the worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create a table in the worksheet
            tab = Table(displayName=f"Presupuestos_{self.entidad}", ref=ws.dimensions)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style

            # Add the table to the worksheet
            ws.add_table(tab)

            # Save the workbook to the response
            wb.save(response)
            return response
        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

class PresupuestosPorClienteProyecto(PresupuestosPorEntidad):
    entidad = 'cliente_proyecto'

class PresupuestosPorProveedor(PresupuestosPorEntidad):
    entidad = 'proveedor'

class GastoPorEntidad(APIView):
    serializer_class = RegistroListSerializer
    queryset = Registro.objects.all()
    entidad = None

    def get(self, request):
        registros = self.get_filtered_registros(request)
        if isinstance(registros, Response):
            return registros

        data = self.serialize_and_process_data(registros)

        if request.query_params.get('export', None) == 'true':
            return self.export_to_excel(data)

        return Response(data)
    
    def serialize_and_process_data(self, registros):
        data = RegistroListSerializer(registros, many=True).data
        
        # Process the data with the annotated dolar_mep_value
        for x in data:
            x['total_gasto_ingreso'] = float(x.get('monto_gasto_ingreso_neto') or 0) + float(x.get('iva_gasto_ingreso') or 0)
            x['monto_gasto_ingreso_neto'] = float(x.get('monto_gasto_ingreso_neto') or 0)
            x['iva_gasto_ingreso'] = float(x.get('iva_gasto_ingreso') or 0)
            x['monto_op_rec'] = float(x.get('monto_op_rec') or 0)
            x['tipo_de_cambio'] = float(x.get('tipo_de_cambio') or 0)
            x['documento'] = ""
            
            # Format the date
            if x.get('fecha_reg'):
                fecha_parts = x['fecha_reg'].split("-")
                x['fecha_reg'] = f"{fecha_parts[2]}/{fecha_parts[1]}/{fecha_parts[0]}"
        
        return data

    def get_filtered_registros(self, request):
        # Filtro por entidad
        if self.entidad:
            entidad_id = request.query_params.get(self.entidad, None)
            if entidad_id is not None:
                registros = Registro.objects.exclude(tipo_reg__in=EXCLUDED_TIPO_REGS).filter(**{self.entidad: entidad_id})
            else:
                return Response(f"Debe especificar un {self.entidad} válido", status=status.HTTP_400_BAD_REQUEST)
        else:
            registros = Registro.objects.exclude(tipo_reg__in=EXCLUDED_TIPO_REGS)

        # Filtro por año y mes
        añomes_max = request.query_params.get('anomes_max', None)
        añomes_min = request.query_params.get('anomes_min', None)
        if añomes_max not in [None, ""]:
            añomes_max = int(añomes_max)
            registros = registros.filter(añomes_imputacion__lte=añomes_max)
        if añomes_min not in [None, ""]:
            añomes_min = int(añomes_min)
            registros = registros.filter(añomes_imputacion__gte=añomes_min)

        # Subquery to get the DolarMEP compra rate for each registro's fecha_reg
        dolar_subquery = DolarMEP.objects.filter(
            fecha=OuterRef('fecha_reg')
        ).values('compra')[:1]  # Only take the first match

        # Annotate the registro queryset with the dolar_mep value
        registros = registros.annotate(
            dolar_mep_value=Coalesce(
                Subquery(dolar_subquery),
                Value(0),  # Default value when no match
                output_field=DecimalField(max_digits=10, decimal_places=2)
            )
        )
        
        return registros

    def export_to_excel(self, data, entidad=None):
        try:
            if not data:
                return Response("No data available to export", status=status.HTTP_400_BAD_REQUEST)
            if not entidad:
                entidad = data[0][self.entidad].replace(" ","_")
            df = pd.DataFrame(data)

            # Create the HttpResponse object with appropriate headers
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Gastos_{entidad}.xlsx'

            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Write DataFrame to the worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create a table in the worksheet
            tab = Table(displayName=f"Gastos_{entidad}", ref=ws.dimensions)

            # Add a default style with striped rows and banded columns
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style

            # Add the table to the worksheet
            ws.add_table(tab)

            # Save the workbook to the response
            wb.save(response)
            return response

        except Exception as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

class DB_Total(APIView):
    serializer_class = RegistroListSerializer
    queryset = Registro.objects.all()
    
    def get(self, request):
        registros = Registro.objects.all()
        temp_view = GastoPorEntidad()  # create instance of GastoPorEntidad
        data = temp_view.serialize_and_process_data(registros)
        return temp_view.export_to_excel(data, 'db_total')

class GastoPorUnidad(GastoPorEntidad):
    entidad = 'unidad_de_negocio'

class GastoPorObra(GastoPorEntidad):
    entidad = 'cliente_proyecto'

class GastoPorProveedor(GastoPorEntidad):
    entidad = 'proveedor'

class CuentasCorrientes(APIView):
    '''
    Nueva cuenta corriente que usa la lógica de cálculo de neto de gestión
    '''
    def get(self, request):
        data = []
        proveedor = request.query_params.get('proveedor', None)
        if proveedor is None:
            cliente_proyecto = request.query_params.get('cliente', None)
            if cliente_proyecto is None:
                return Response({'detail': 'Debe especificar un proveedor o un cliente'}, status=status.HTTP_400_BAD_REQUEST)
            else: 
                registros = Registro.objects.filter(cliente_proyecto__id=int(cliente_proyecto), tipo_reg__in=['FCV', 'REC', 'RECFC', 'ISF', 'RETS'], activo=True).exclude(imputacion__imputacion="Diferencia de cambio").order_by('fecha_reg')
        else:
            registros = Registro.objects.filter(proveedor__id=int(proveedor), activo=True).exclude(imputacion__imputacion="Diferencia de cambio").order_by('fecha_reg')
        monto_op_rec = registros.aggregate(saldo=Sum('monto_op_rec'))['saldo'] or 0
        monto_gasto_ingreso_neto = registros.aggregate(saldo=Sum('monto_gasto_ingreso_neto'))['saldo'] or 0
        iva_gasto_ingreso = registros.aggregate(saldo=Sum('iva_gasto_ingreso'))['saldo'] or 0
        saldo = -monto_op_rec + monto_gasto_ingreso_neto + iva_gasto_ingreso
        data.append({
            'persona': Persona.objects.get(id=proveedor).nombre_fantasia if proveedor else ClienteProyecto.objects.get(id=cliente_proyecto).cliente_proyecto,
            'registros': RegistroCC(registros, many=True).data,
            'saldo': saldo
        })
        return Response(data)

class MovimientosDeCaja(APIView):
    def get(self, request):
        caja = request.query_params.get('caja', None)
        fecha_max = request.query_params.get('fecha_max', None)
        fecha_min = request.query_params.get('fecha_min', None)

        if caja is None or fecha_max is None or fecha_min is None:
            return Response({'detail': 'Debe especificar una caja y un rango de fechas'}, status=status.HTTP_400_BAD_REQUEST)

        registros = Registro.objects.filter(
            caja__id=int(caja),
            fecha_reg__range=[fecha_min, fecha_max]
        ).order_by('fecha_reg', 'id')

        # Calcular el saldo inicial
        saldo_inicial = Registro.objects.filter(
            caja__id=int(caja),
            fecha_reg__lt=fecha_min
        ).aggregate(saldo_inicial=Sum('monto_op_rec'))['saldo_inicial'] or 0

        # Calcular el saldo acumulado para cada registro
        registros = registros.annotate(
            saldo_parcial=F('monto_op_rec')
        ).annotate(
            saldo_acumulado=Window(
                expression=Sum(F('saldo_parcial')),
                order_by=(F('fecha_reg').asc(), F('id').asc())
            )
        )

        data = RegistroSerializer(registros, many=True).data

        # Agregar el saldo inicial al primer registro
        if data:
            for registro in data:
                registro['saldo_inicial'] = saldo_inicial
                registro['saldo_acumulado'] = Decimal(registro['saldo_acumulado']) + saldo_inicial

        # Calcular el saldo final (último saldo_acumulado)
        saldo_final = saldo_inicial
        if data:
            saldo_final = data[-1]['saldo_acumulado']

        return Response({
            'saldo_inicial': saldo_inicial,
            'saldo_final': saldo_final,
            'registros': data
        })